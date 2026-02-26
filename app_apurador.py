"""
Apurador de Investimentos - Aplicação Web Streamlit

Interface web moderna para apuração de investimentos em promoções.
Permite upload de planilhas e download de resultados diretamente pelo navegador.

Autor: Nivea Project
Data: 2026-02-09
"""

import streamlit as st
import pandas as pd
import io
import os
from io import BytesIO
from datetime import datetime
import openpyxl
import openpyxl.styles


# Configuração da página
st.set_page_config(
    page_title="Apurador de Investimentos",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS customizado para melhorar a aparência
st.markdown("""
<style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #2c3e50;
        text-align: center;
        margin-bottom: 1rem;
    }
    .sub-header {
        font-size: 1.2rem;
        color: #7f8c8d;
        text-align: center;
        margin-bottom: 2rem;
    }
    .stAlert {
        margin-top: 1rem;
    }
    .upload-section {
        background-color: #f8f9fa;
        padding: 1.5rem;
        border-radius: 10px;
        margin-bottom: 1rem;
    }
    .success-box {
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        color: #155724;
        padding: 1rem;
        border-radius: 5px;
        margin: 1rem 0;
    }
    .info-box {
        background-color: #d1ecf1;
        border: 1px solid #bee5eb;
        color: #0c5460;
        padding: 1rem;
        border-radius: 5px;
        margin: 1rem 0;
    }
</style>
""", unsafe_allow_html=True)


def validar_colunas_preco_final(df):
    """Valida se a planilha de preço final tem as colunas necessárias"""
    # Aceita tanto "EAN" quanto "COD BARRAS"
    tem_identificador = 'EAN' in df.columns or 'COD BARRAS' in df.columns
    
    if not tem_identificador:
        return False, "A planilha deve conter a coluna 'EAN' ou 'COD BARRAS'"
    
    # Padronizar coluna para EAN
    if 'COD BARRAS' in df.columns and 'EAN' not in df.columns:
        df.rename(columns={'COD BARRAS': 'EAN'}, inplace=True)
    
    return True, "Planilha válida"


def validar_colunas_orcamento(df):
    """Valida se a planilha de orçamento tem as colunas necessárias"""
    colunas_necessarias = ['EAN', 'VALOR SKU PAGO', 'QUANTIDADE']
    colunas_faltando = [col for col in colunas_necessarias if col not in df.columns]
    
    if colunas_faltando:
        return False, f"Colunas faltando: {', '.join(colunas_faltando)}"
    
    return True, "Planilha válida"


def limpar_valor_monetario(valor):
    """
    Limpa valores monetários em formato brasileiro ou americano para conversão numérica
    Exemplos: 
    - 'R$ 1.234,56' -> 1234.56 (BR: ponto=milhar, vírgula=decimal)
    - '1234,56' -> 1234.56 (BR: vírgula=decimal)
    - '1234.56' -> 1234.56 (US: ponto=decimal)
    - '20,68' -> 20.68 (BR: vírgula=decimal)
    - '20.68' -> 20.68 (US: ponto=decimal, já está correto)
    """
    if pd.isna(valor):
        return None
    
    # Converter para string
    valor_str = str(valor).strip()
    
    # Remover símbolos de moeda e espaços
    valor_str = valor_str.replace('R$', '').replace('r$', '').replace('$', '').strip()
    
    # Detectar formato:
    # Se tem vírgula E ponto, verificar qual vem por último
    tem_virgula = ',' in valor_str
    tem_ponto = '.' in valor_str
    
    if tem_virgula and tem_ponto:
        # Ambos presentes: o último é o separador decimal
        pos_virgula = valor_str.rfind(',')
        pos_ponto = valor_str.rfind('.')
        
        if pos_virgula > pos_ponto:
            # Formato BR: 1.234,56 (ponto=milhar, vírgula=decimal)
            valor_str = valor_str.replace('.', '').replace(',', '.')
        else:
            # Formato US: 1,234.56 (vírgula=milhar, ponto=decimal)
            valor_str = valor_str.replace(',', '')
    elif tem_virgula:
        # Só vírgula: formato BR - vírgula é decimal
        valor_str = valor_str.replace(',', '.')
    # Se só tem ponto ou nenhum: já está no formato correto (US ou inteiro)
    
    # Converter para float
    try:
        return float(valor_str)
    except:
        return None



def processar_dados(df_preco_final, orcamentos_dict):
    """Processa os dados e calcula investimentos"""
    # Copiar dataframe de preço final
    df_resultado = df_preco_final.copy()
    
    # Converter EAN para string e remover espaços
    df_resultado['EAN'] = df_resultado['EAN'].astype(str).str.strip()
    
    # Verificar se existe coluna de valor negociado (case-insensitive)
    coluna_valor_negociado = None
    colunas_upper = {col.upper(): col for col in df_resultado.columns}
    
    for possivel_nome in ['VALOR NEGOCIADO REDE', 'VALOR NEGOCIADO', 'PRECO NEGOCIADO', 'PREÇO NEGOCIADO']:
        if possivel_nome in colunas_upper:
            coluna_valor_negociado = colunas_upper[possivel_nome]
            break
    
    if coluna_valor_negociado is None:
        st.error("❌ Não foi encontrada coluna de valor negociado na planilha de Preço Final")
        st.info("💡 Colunas aceitas: 'VALOR NEGOCIADO REDE', 'VALOR NEGOCIADO', 'PRECO NEGOCIADO'")
        return None, None
    
    # Converter valor negociado para numérico antes das comparações
    df_resultado[coluna_valor_negociado] = df_resultado[coluna_valor_negociado].apply(limpar_valor_monetario)
    
    # Coletar todos os EANs presentes nos orçamentos
    eans_orcamentos = set()
    for df_orc in orcamentos_dict.values():
        eans_str = df_orc['EAN'].astype(str).str.strip()
        eans_orcamentos.update(eans_str.tolist())
    
    # Filtrar apenas produtos do Preço Final que estão em algum orçamento
    eans_resultado = df_resultado['EAN'].astype(str).str.strip()
    df_no_orcamento = df_resultado[eans_resultado.isin(eans_orcamentos)]
    
    # Detectar produtos sem preço entre os que estão nos orçamentos
    sem_preco = df_no_orcamento[
        df_no_orcamento[coluna_valor_negociado].isna() | (df_no_orcamento[coluna_valor_negociado] == 0)
    ]
    
    if not sem_preco.empty:
        col_produto = next((c for c in df_resultado.columns if 'produto' in c.lower() or 'descri' in c.lower()), None)
        col_ean = 'EAN' if 'EAN' in df_resultado.columns else None
        
        st.error(f"❌ **{len(sem_preco)} produto(s) presentes no orçamento estão sem preço negociado (zero ou vazio)**. Corrija antes de continuar.")
        
        with st.expander("📋 Ver lista de produtos sem preço", expanded=True):
            colunas_exibir = []
            if col_ean:
                colunas_exibir.append(col_ean)
            if col_produto:
                colunas_exibir.append(col_produto)
            colunas_exibir.append(coluna_valor_negociado)
            
            if colunas_exibir:
                st.dataframe(sem_preco[colunas_exibir].reset_index(drop=True), use_container_width=True)
            else:
                st.dataframe(sem_preco.reset_index(drop=True), use_container_width=True)
        
        return None, None
    

    for nome, df_orc in orcamentos_dict.items():
        # Converter EAN para string
        df_orc['EAN'] = df_orc['EAN'].astype(str).str.strip()
        
        # Preparar dados do orçamento
        df_orc_temp = df_orc[['EAN', 'VALOR SKU PAGO', 'QUANTIDADE']].copy()
        
        # Debug: mostrar alguns EANs do orçamento
        st.caption(f"🔍 {nome} - Primeiros EANs: {df_orc_temp['EAN'].head(3).tolist()}")
        
        # Debug: mostrar valores originais antes da limpeza
        st.caption(f"🔍 {nome} - Exemplo VALOR SKU PAGO original: {df_orc_temp['VALOR SKU PAGO'].head(3).tolist()}")
        
        # Limpar e converter valores para numérico
        df_orc_temp['VALOR SKU PAGO'] = df_orc_temp['VALOR SKU PAGO'].apply(limpar_valor_monetario)
        df_orc_temp['QUANTIDADE'] = pd.to_numeric(df_orc_temp['QUANTIDADE'], errors='coerce')
        
        # Debug: mostrar quantos valores válidos temos
        valores_validos = df_orc_temp['VALOR SKU PAGO'].notna().sum()
        qtd_validas = df_orc_temp['QUANTIDADE'].notna().sum()
        st.caption(f"📊 {nome}: {valores_validos} valores SKU válidos, {qtd_validas} quantidades válidas (de {len(df_orc_temp)} linhas)")
        
        df_orc_temp = df_orc_temp.rename(columns={
            'VALOR SKU PAGO': f'{nome}_VALOR_SKU_PAGO',
            'QUANTIDADE': f'{nome}_QUANTIDADE'
        })
        
        # Merge por EAN (LEFT JOIN)
        antes_merge = len(df_resultado)
        df_resultado = df_resultado.merge(
            df_orc_temp,
            on='EAN',
            how='left'
        )
        
        # Debug: verificar quantos matches foram feitos
        matches = df_resultado[f'{nome}_QUANTIDADE'].notna().sum()
        st.caption(f"✅ {nome}: {matches} produtos encontrados no Preço Final (de {antes_merge} produtos)")
        
        if matches == 0:
            st.warning(f"⚠️ Nenhum produto de '{nome}' foi encontrado no Preço Final. Verifique se os EANs são iguais!")
    
    # Debug: mostrar alguns EANs do Preço Final
    st.caption(f"🔍 Preço Final - Primeiros EANs: {df_resultado['EAN'].head(3).tolist()}")
    
    # Calcular investimentos e valores para cada orçamento
    for nome in orcamentos_dict.keys():
        
        # Calcular Investimento Total e Valor de Pedido
        df_resultado[f'{nome}_INVESTIMENTO_TOTAL'] = (
            (df_resultado[f'{nome}_VALOR_SKU_PAGO'] - df_resultado[coluna_valor_negociado]) * 
            df_resultado[f'{nome}_QUANTIDADE']
        )
        
        df_resultado[f'{nome}_VALOR_PEDIDO_TOTAL'] = (
            df_resultado[f'{nome}_VALOR_SKU_PAGO'] * df_resultado[f'{nome}_QUANTIDADE']
        )
    
    # Criar colunas totalizadoras (soma de todos os orçamentos)
    colunas_investimento = [f'{nome}_INVESTIMENTO_TOTAL' for nome in orcamentos_dict.keys()]
    colunas_valor_pedido = [f'{nome}_VALOR_PEDIDO_TOTAL' for nome in orcamentos_dict.keys()]
    
    # Somar todas as colunas de investimento (ignorando NaN)
    df_resultado['INVESTIMENTO_TOTAL_GERAL'] = df_resultado[colunas_investimento].sum(axis=1, skipna=True)
    
    # Somar todas as colunas de valor de pedido (ignorando NaN)
    df_resultado['VALOR_PEDIDO_TOTAL_GERAL'] = df_resultado[colunas_valor_pedido].sum(axis=1, skipna=True)
    
    # Coletar estatísticas ANTES de renomear/remover colunas (para exibir no Streamlit)
    estatisticas = {}
    for nome in orcamentos_dict.keys():
        produtos_encontrados = df_resultado[f'{nome}_QUANTIDADE'].notna().sum()
        total_produtos = len(df_resultado)
        estatisticas[nome] = {
            'encontrados': produtos_encontrados,
            'total': total_produtos
        }
    
    # Remover colunas individuais de investimento e valor de pedido (manter só os totais)
    colunas_para_remover = colunas_investimento + colunas_valor_pedido
    df_resultado = df_resultado.drop(columns=colunas_para_remover)
    
    # Renomear colunas para nomes mais amigáveis
    renomeacoes = {}
    
    # Renomear colunas dos orçamentos dinamicamente
    nomes_orcamentos = list(orcamentos_dict.keys())
    for idx, nome in enumerate(nomes_orcamentos, 1):
        renomeacoes[f'{nome}_VALOR_SKU_PAGO'] = f'Preço venda loja {idx}'
        renomeacoes[f'{nome}_QUANTIDADE'] = f'Qtd venda loja {idx}'
    
    # Renomear colunas totais
    renomeacoes['INVESTIMENTO_TOTAL_GERAL'] = 'Verba Total'
    renomeacoes['VALOR_PEDIDO_TOTAL_GERAL'] = 'TT.Pedido'
    
    df_resultado = df_resultado.rename(columns=renomeacoes)
    
    # Calcular % Investimento (Verba Total / TT.Pedido * 100)
    df_resultado['% Investimento'] = (
        (df_resultado['Verba Total'] / df_resultado['TT.Pedido']) * 100
    ).round(2)  # Arredondar para 2 casas decimais
    
    # Substituir inf e NaN por 0 (quando TT.Pedido = 0)
    df_resultado['% Investimento'] = df_resultado['% Investimento'].replace([float('inf'), -float('inf')], 0).fillna(0)
    
    return df_resultado, estatisticas


def converter_df_para_excel(df, nome_rede=""):
    """Converte DataFrame para Excel em memória com formatação e resumo"""
    output = io.BytesIO()
    
    # Calcular totais para o resumo
    total_verba = df['Verba Total'].sum()
    total_pedido = df['TT.Pedido'].sum()
    percentual_investimento = (total_verba / total_pedido * 100) if total_pedido > 0 else 0
    
    # Criar título com nome da rede e data
    data_atual = datetime.now().strftime('%d/%m/%Y')
    titulo = f"RESUMO - {nome_rede} - {data_atual}" if nome_rede else f"RESUMO - {data_atual}"
    
    # Criar writer com engine openpyxl
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Criar DataFrame vazio para o resumo (linhas 1-4)
        # Linha 1: Título
        # Linha 2: Cabeçalhos do resumo
        # Linha 3: Valores do resumo
        # Linha 4: Vazia (separador)
        # Linha 5+: Dados
        
        # Escrever dados principais a partir da linha 5 (startrow=4)
        df.to_excel(writer, index=False, sheet_name='Apuração', startrow=4)
        
        # Obter worksheet para aplicar formatação
        worksheet = writer.sheets['Apuração']
        
        # Adicionar título do resumo na linha 1
        worksheet['A1'] = titulo
        worksheet['A1'].font = openpyxl.styles.Font(size=11, bold=True)
        
        # Adicionar cabeçalhos do resumo na linha 2
        worksheet['A2'] = 'Verba Total'
        worksheet['B2'] = 'TT.Pedido'
        worksheet['C2'] = '% Investimento'
        
        # Aplicar negrito nos cabeçalhos
        for cell in ['A2', 'B2', 'C2']:
            worksheet[cell].font = openpyxl.styles.Font(bold=True)
        
        # Adicionar valores do resumo na linha 3
        worksheet['A3'] = total_verba
        worksheet['B3'] = total_pedido
        worksheet['C3'] = percentual_investimento
        
        # Formatar valores do resumo
        worksheet['A3'].number_format = 'R$ #,##0.00'
        worksheet['B3'].number_format = 'R$ #,##0.00'
        worksheet['C3'].number_format = '0.00"%"'
        
        # Ajustar largura das colunas
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Aplicar formatação de moeda R$ e percentual nos dados principais
        # Identificar colunas que precisam de formatação
        colunas_moeda = ['Valor Negociado REDE', 'Verba Total', 'TT.Pedido']
        colunas_moeda_dinamicas = [col for col in df.columns if 'Preço venda loja' in col]
        todas_colunas_moeda = colunas_moeda + colunas_moeda_dinamicas
        
        coluna_percentual = '% Investimento'
        
        # Obter índices das colunas
        indices_moeda = []
        indice_percentual = None
        
        for idx, col in enumerate(df.columns, 1):  # Excel columns são 1-indexed
            if col in todas_colunas_moeda:
                indices_moeda.append(idx)
            elif col == coluna_percentual:
                indice_percentual = idx
        
        # Aplicar formatação (começar da linha 6 pois linha 5 é cabeçalho dos dados)
        for row in range(6, len(df) + 6):
            # Formatar colunas de moeda
            for col_idx in indices_moeda:
                cell = worksheet.cell(row=row, column=col_idx)
                cell.number_format = 'R$ #,##0.00'
            
            # Formatar coluna de percentual
            if indice_percentual:
                cell = worksheet.cell(row=row, column=indice_percentual)
                cell.number_format = '0.00"%"'
        
        # Aplicar cores de fundo nas células
        from openpyxl.styles import PatternFill
        
        # Definir as cores
        cor_cabecalho = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')  # Azul escuro
        cor_dados_azul = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')  # Azul claro
        cor_dados_verde = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')  # Verde claro
        cor_orcamentos = PatternFill(start_color='FEF2CB', end_color='FEF2CB', fill_type='solid')  # Bege claro
        cor_resumo_preto = PatternFill(start_color='000000', end_color='000000', fill_type='solid')  # Preto
        cor_resumo_amarelo = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Amarelo
        cor_resumo_verde = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')  # Verde
        
        # Aplicar cores no resumo (A2:C2) - Fundo preto, fonte branca
        for col in range(1, 4):  # Colunas A, B, C
            cell = worksheet.cell(row=2, column=col)
            cell.fill = cor_resumo_preto
            cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        
        # Aplicar cores no resumo linha 3
        # A3 - Amarelo
        worksheet['A3'].fill = cor_resumo_amarelo
        # B3 - Verde
        worksheet['B3'].fill = cor_resumo_verde
        # C3 - Amarelo
        worksheet['C3'].fill = cor_resumo_amarelo
        
        # Aplicar cor no cabeçalho (A5:E5)
        for col in range(1, 6):  # Colunas A até E (1 até 5)
            cell = worksheet.cell(row=5, column=col)
            cell.fill = cor_cabecalho
            # Aplicar texto branco no cabeçalho para contraste
            cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        
        # Aplicar cor azul claro nos dados (A6:E179)
        for row in range(6, 180):  # Linhas 6 até 179
            for col in range(1, 6):  # Colunas A até E
                cell = worksheet.cell(row=row, column=col)
                cell.fill = cor_dados_azul
        
        # Aplicar cor verde claro na coluna F (F5:F179)
        for row in range(5, 180):  # Linhas 5 até 179
            cell = worksheet.cell(row=row, column=6)  # Coluna F
            cell.fill = cor_dados_verde
            if row == 5:  # Aplicar negrito e branco no cabeçalho
                cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
                cell.fill = cor_cabecalho
        
        # Identificar e aplicar cor nas colunas de orçamentos (qtd e preço venda)
        # Buscar por colunas que contêm "Preço venda loja" ou "Qtd venda loja"
        colunas_orcamento = []
        for idx, col_name in enumerate(df.columns, 1):
            col_lower = str(col_name).lower()
            # Verificar se é coluna de orçamento (preço venda ou qtd venda)
            if 'preço venda loja' in col_lower or 'qtd venda loja' in col_lower:
                colunas_orcamento.append(idx)
        
        # Aplicar cor bege claro nas colunas de orçamentos
        for col_idx in colunas_orcamento:
            for row in range(5, 180):  # Linhas 5 até 179
                cell = worksheet.cell(row=row, column=col_idx)
                cell.fill = cor_orcamentos
                if row == 5:  # Manter negrito e branco no cabeçalho
                    cell.font = openpyxl.styles.Font(bold=True, color='FFFFFF')
                    cell.fill = cor_cabecalho
    
    output.seek(0)
    return output.getvalue()


def main():
    """Função principal da aplicação"""
    
    # Cabeçalho
    st.markdown('<div class="main-header">📊 Apurador de Investimentos</div>', unsafe_allow_html=True)
    st.markdown('<div class="sub-header">Calcule investimentos em promoções de forma rápida e eficiente</div>', unsafe_allow_html=True)
    
    # Botões de ajuda e template
    col_help1, col_help2, col_help3, col_help4 = st.columns([1, 1, 1, 1])
    
    with col_help1:
        st.link_button(
            "🎥 Tutorial em Vídeo",
            "https://www.loom.com/share/a50e2261f1e84bd6a223160e7b69ad1e",
            help="Assista ao tutorial completo de como usar o sistema",
            use_container_width=True
        )
    
    with col_help2:
        st.link_button(
            "💬 Suporte WhatsApp",
            "https://wa.me/5534999079685?text=Olá! Preciso de ajuda com o Apurador de Investimentos",
            help="Entre em contato em caso de dúvidas",
            use_container_width=True
        )
    
    with col_help3:
        # Carregar o arquivo template Nivea para download
        try:
            # Usar caminho absoluto baseado na localização do script
            template_path = os.path.join(os.path.dirname(__file__), "SIMULADOR_NIVEA_2026.xlsx")
            with open(template_path, "rb") as file:
                template_data = file.read()
            
            st.download_button(
                "📋 modelo padrão Nivea",
                data=template_data,
                file_name="SIMULADOR_NIVEA_2026.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Baixe a planilha modelo padrão Nivea",
                use_container_width=True
            )
        except FileNotFoundError:
            st.warning("⚠️ Planilha modelo Nivea não encontrada")
    
    with col_help4:
        # Carregar o arquivo template Reckitt para download
        try:
            template_path = os.path.join(os.path.dirname(__file__), "SIMULADOR_RECKITT 2026.xlsx")
            with open(template_path, "rb") as file:
                template_data = file.read()
            
            st.download_button(
                "📋 modelo padrão Reckitt",
                data=template_data,
                file_name="SIMULADOR_RECKITT 2026.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Baixe a planilha modelo padrão Reckitt",
                use_container_width=True,
                key="download_reckitt"
            )
        except FileNotFoundError:
            st.warning("⚠️ Planilha modelo Reckitt não encontrada")
    
    st.markdown("---")
    
    # Barra lateral com instruções
    with st.sidebar:
        st.header("📋 Instruções")
        st.markdown("""
        ### Como usar:
        
        1️⃣ **Preço Final**  
        Upload da planilha com:
        - EAN ou COD BARRAS
        - Preço negociado
        
        2️⃣ **Orçamentos**  
        Upload de 1 ou mais planilhas com:
        - EAN
        - VALOR SKU PAGO
        - QUANTIDADE
        
        ⚠️ **Importante**: Especifique em qual linha estão os cabeçalhos (padrão: linha 8)
        
        3️⃣ **Processar**  
        Clique no botão para calcular
        
        4️⃣ **Download**  
        Baixe o resultado em Excel
        """)
        
        st.divider()
        
        st.header("📊 Cálculos")
        st.markdown("""
        **Investimento Total:**
        ```
        (Valor SKU - Valor Negociado) 
        × Quantidade
        ```
        
        **Valor de Pedido:**
        ```
        Valor SKU × Quantidade
        ```
        """)
    
    # Inicializar session state
    if 'df_preco_final' not in st.session_state:
        st.session_state.df_preco_final = None
    if 'orcamentos_dict' not in st.session_state:
        st.session_state.orcamentos_dict = {}
    if 'df_resultado' not in st.session_state:
        st.session_state.df_resultado = None
    
    # Container para upload de Preço Final
    st.subheader("1️⃣ Planilha de Preço Final")
    
    col1, col2 = st.columns([2, 1])
    
    with col1:
        arquivo_preco = st.file_uploader(
            "Selecione a planilha de Preço Final",
            type=['xlsx', 'xls'],
            key="upload_preco",
            help="Upload do arquivo Excel com EAN/COD BARRAS e preço negociado"
        )
    
    with col2:
        nome_rede_input = st.text_input(
            "Nome da Rede",
            placeholder="Ex: REDE ABC",
            help="Nome da rede para incluir no relatório"
        )
    
    if arquivo_preco:
        try:
            df_preco = pd.read_excel(arquivo_preco)
            
            valido, mensagem = validar_colunas_preco_final(df_preco)
            
            if valido:
                st.session_state.df_preco_final = df_preco
                st.session_state.nome_rede = nome_rede_input if nome_rede_input else "[REDE]"
                st.success(f"✅ Preço Final carregado com sucesso!")
                with col2:
                    st.metric("📦 Produtos", len(df_preco))
                
                with st.expander("👁️ Visualizar dados carregados"):
                    st.dataframe(df_preco.head(10), use_container_width=True)
            else:
                st.error(f"❌ {mensagem}")
                st.session_state.df_preco_final = None
                
        except Exception as e:
            st.error(f"❌ Erro ao ler arquivo: {str(e)}")
            st.session_state.df_preco_final = None
    
    st.divider()
    
    # Container para upload de Orçamentos
    st.subheader("2️⃣ Planilhas de Orçamento")
    
    st.info("💡 As planilhas de orçamento devem ter os cabeçalhos na **linha 8** e conter as colunas: EAN, VALOR SKU PAGO, QUANTIDADE")
    
    arquivos_orcamento = st.file_uploader(
        "Selecione uma ou mais planilhas de orçamento",
        type=['xlsx', 'xls'],
        accept_multiple_files=True,
        key="upload_orcamentos"
    )
    
    if arquivos_orcamento:
        st.session_state.orcamentos_dict = {}
        
        for arquivo in arquivos_orcamento:
            try:
                # Ler Excel com cabeçalho na linha 8 (header=7 porque é 0-indexed)
                df_temp = pd.read_excel(arquivo, header=7)
                
                valido, mensagem = validar_colunas_orcamento(df_temp)
                
                if valido:
                    nome_orcamento = arquivo.name.replace('.xlsx', '').replace('.xls', '')
                    st.session_state.orcamentos_dict[nome_orcamento] = df_temp
                    
                    col1, col2, col3 = st.columns([2, 1, 1])
                    with col1:
                        st.success(f"✅ {arquivo.name}")
                    with col2:
                        st.metric("📦 Produtos", len(df_temp))
                    with col3:
                        with st.expander("👁️ Ver"):
                            st.dataframe(df_temp.head(5), use_container_width=True)
                else:
                    st.error(f"❌ {arquivo.name}: {mensagem}")
                    
            except Exception as e:
                st.error(f"❌ Erro ao ler {arquivo.name}: {str(e)}")
    
    st.divider()
    
    # Botão de processamento
    st.subheader("3️⃣ Processar Dados")
    
    col1, col2, col3 = st.columns([1, 2, 1])
    
    with col2:
        processar_btn = st.button(
            "🚀 Processar e Calcular Investimentos",
            type="primary",
            use_container_width=True,
            disabled=(st.session_state.df_preco_final is None or not st.session_state.orcamentos_dict)
        )
    
    if processar_btn:
        with st.spinner("⏳ Processando dados..."):
            resultado = processar_dados(
                st.session_state.df_preco_final,
                st.session_state.orcamentos_dict
            )
            
            if resultado is None or resultado[0] is None:
                pass  # erros já exibidos dentro da função
            else:
                df_resultado, estatisticas = resultado
            
            if resultado is not None and resultado[0] is not None:
                st.session_state.df_resultado = df_resultado
                
                st.success("✅ Processamento concluído com sucesso!")
                
                # Estatísticas
                st.subheader("📈 Resumo do Processamento")
                
                cols = st.columns(len(estatisticas) + 1)
                
                with cols[0]:
                    st.metric(
                        "Total de Produtos",
                        len(df_resultado),
                        help="Total de produtos na planilha de Preço Final"
                    )
                
                for idx, (nome, stats) in enumerate(estatisticas.items(), 1):
                    with cols[idx]:
                        st.metric(
                            f"📦 {nome}",
                            f"{stats['encontrados']}/{stats['total']}",
                            help=f"Produtos encontrados neste orçamento"
                        )
    
    # Seção de download e visualização
    if st.session_state.df_resultado is not None:
        st.divider()
        st.subheader("4️⃣ Resultado")
        
        # Visualização dos dados
        with st.expander("👁️ Visualizar Resultado Completo", expanded=False):
            st.dataframe(st.session_state.df_resultado, use_container_width=True)
        
        # Download
        nome_arquivo = f"Apuracao_Investimentos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        excel_data = converter_df_para_excel(
            st.session_state.df_resultado,
            st.session_state.get('nome_rede', '')
        )
        
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.download_button(
                label="📥 Download Resultado (Excel)",
                data=excel_data,
                file_name=nome_arquivo,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True
            )
        
        st.info("💡 O arquivo Excel contém todos os dados do Preço Final mais as colunas:\n" +
                "• Preço venda loja 1, 2, etc.\n" +
                "• Qtd venda loja 1, 2, etc.\n" +
                "• Verba Total (investimento total)\n" +
                "• TT.Pedido (valor total de pedidos)")
    
    # Rodapé
    st.divider()
    st.markdown(
        "<div style='text-align: center; color: #7f8c8d; font-size: 0.9rem;'>"
        "Apurador de Investimentos v2.0 | Streamlit Web App | 2026"
        "</div>",
        unsafe_allow_html=True
    )


if __name__ == "__main__":
    main()
